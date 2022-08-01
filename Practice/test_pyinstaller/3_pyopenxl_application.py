import openpyxl
import time

file_name = '3_some_excel_data.xlsx'
excelFile = openpyxl.load_workbook(file_name)
sheet = excelFile.active
data = []
temp = []
for i in range(10):
    for j in range(10):
        temp.append(sheet.cell(row=i+1, column=j+1).value)
    data.append(temp)
    temp = []

print(data)

for i in range(10):
    for j in range(10):
        if type(data[i][j]) == int:
            sheet.cell(row=i+1, column=j+1, value = int(data[i][j])+10)

excelFile.save(file_name)

user_input = None
while user_input != 'y':
    user_input = input('terminate? Y/N')
